from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.db import transaction
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.urls import reverse_lazy
from django.utils.translation import gettext_lazy as _
from django.views import View
from django.views.generic import FormView
from openpyxl import load_workbook

from quizz.forms.management import ImportQuestionsForm
from quizz.models import (
    QuestionLocale,
    Contest,
    QUESTION_OPEN,
    QUESTION_LINKED,
    QUESTION_MCQ,
    Tag,
    Question,
)


class QuestionsImportView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    template_name = "management/question-import.html"
    form_class = ImportQuestionsForm
    permission_required = "quizz.add_question"

    @transaction.atomic
    def form_valid(self, form):
        wb = load_workbook(self.request.FILES["spreadsheet"])
        ws = wb.active

        header_found = False
        values_started = False

        created_questions = []
        errored_questions = []

        locales = {}
        sources = {}
        tags = {}

        for row in ws.iter_rows(values_only=True):
            len_row = len(row)

            # We try to detect and skip an eventual header: the first row with
            # label-style values. If we miss it, it's not the end of the world
            # as such values will probably be invalid anyway (and users may still
            # delete the invalid question afterward).
            if not values_started and not header_found and row:
                normalized_row = [str(cell).lower() for cell in row]

                if (
                    (
                        len_row > 0
                        and (
                            "lang" in normalized_row[0]
                            or "language" in normalized_row[0]
                            or "locale" in normalized_row[0]
                        )
                    )
                    and (len_row > 1 and "question" in normalized_row[1])
                    and (
                        len_row > 2
                        and "proposition" in normalized_row[2]
                        or "proposed" in normalized_row[2]
                    )
                    and (
                        len_row > 3
                        and (
                            "answer" in normalized_row[3]
                            or "réponse" in normalized_row[3]
                            or "réponse" in normalized_row[3]
                        )
                    )
                    and (len_row > 4 and "comment" in normalized_row[4])
                    and (len_row > 5 and "tag" in normalized_row[5])
                    and (len_row > 6 and "origin" in normalized_row[6])
                    and (len_row > 7 and "difficult" in normalized_row[7])
                ):
                    header_found = True
                    values_started = True
                    continue

            if row:
                values_started = True

            # We store the raw content of the future question in a dict
            raw_question = {
                "locale": str(row[0]).strip() if len_row > 0 and row[0] else None,
                "question": str(row[1]).strip() if len_row > 1 and row[1] else None,
                "proposed_answers": str(row[2]).splitlines()
                if len_row > 2 and row[2]
                else None,
                "answers": str(row[3]).splitlines() if len_row > 3 and row[3] else None,
                "comment": str(row[4]).strip() if len_row > 4 and row[4] else None,
                "tags": str(row[5]).splitlines() if len_row > 5 and row[5] else None,
                "source": str(row[6]).strip() if len_row > 6 and row[6] else None,
                "difficulty": str(row[7]).strip() if len_row > 7 and row[7] else None,
            }

            # If the question is already obviously invalid (missing required
            # fields), we skip it.
            if not raw_question["question"] or (
                not raw_question["proposed_answers"] and not raw_question["answers"]
            ):
                errored_questions.append((raw_question, "missing_fields"))
                continue

            # Locale
            if raw_question["locale"].lower() in locales:
                locale = locales[raw_question["locale"].lower()]
            else:
                locale = QuestionLocale.objects.filter(
                    name__istartswith=raw_question["locale"]
                ).first()
                if not locale:
                    locale = QuestionLocale.objects.filter(
                        code__istartswith=raw_question["locale"]
                    ).first()
                if not locale:
                    errored_questions.append((raw_question, "locale"))
                    continue

                locales[raw_question["locale"].lower()] = locale

            # Difficulty
            if raw_question["difficulty"]:
                normalized_difficulty = raw_question["difficulty"].lower()
                if normalized_difficulty in ["1", "e", "easy"]:
                    difficulty = 1
                elif normalized_difficulty in ["2", "m", "medium"]:
                    difficulty = 2
                elif normalized_difficulty in ["3", "h", "hard"]:
                    difficulty = 3
                else:
                    difficulty = int(form.cleaned_data["default_difficulty"])
            else:
                difficulty = int(form.cleaned_data["default_difficulty"])

            # Source
            # We try to find a corresponding contest (skipping punctuation if needed);
            # if nothing can be found, we create a new one.
            source = None
            if raw_question["source"]:
                normalized_source = raw_question["source"].lower()
                if normalized_source in sources:
                    source = sources[normalized_source]
                else:
                    source = Contest.objects.filter(
                        name__iexact=normalized_source
                    ).first()
                    if not source:
                        no_pct_source = (
                            normalized_source.replace(",", "")
                            .replace(".", "")
                            .replace(";", "")
                            .replace(":", "")
                        )
                        source = Contest.objects.filter(
                            name__iexact=no_pct_source
                        ).first()
                        if not source:
                            source = Contest(name=raw_question["source"])
                            source.save()
                    sources[normalized_source] = source

            # Question type
            if not raw_question["proposed_answers"]:
                question_type = QUESTION_OPEN
            elif not raw_question["answers"]:
                question_type = (
                    QUESTION_LINKED
                    if all(
                        ["-->" in answer for answer in raw_question["proposed_answers"]]
                    )
                    else QUESTION_MCQ
                )
            else:
                question_type = QUESTION_MCQ

            # Answers
            has_open_choice = False
            open_choice = None

            if question_type == QUESTION_OPEN:
                answers = "\n".join(
                    [answer.strip() for answer in raw_question["answers"]]
                )

            elif question_type == QUESTION_MCQ:
                answers = []

                normalized_proposed_answers = [
                    answer.lower().strip()
                    for answer in raw_question["proposed_answers"]
                ]
                normalized_answers = [
                    answer.lower().strip() for answer in raw_question["answers"]
                ]

                for answer in raw_question["proposed_answers"]:
                    if answer.lower() in [
                        "other:",
                        "other :",
                        "autre:",
                        "autre :",  # space [SP]
                        "autre :",  # no-break space [NBSP]
                        "autre :",  # narrow no-break space [NNBSP]
                        "autres:",
                        "autres :",  # space [SP]
                        "autres :",  # no-break space [NBSP]
                        "autres :",  # narrow no-break space [NNBSP]
                    ]:
                        has_open_choice = True
                    else:
                        answers.append(
                            {
                                "answer": answer.strip(),
                                "is_correct": answer.lower().strip()
                                in normalized_answers,
                            }
                        )

                # We process unknown answers here (i.e. answers in the “Answers”
                # list but not in the “Proposed answers” one).
                # If we have an open answer, these will be the answer, and they
                # will be concatenated with new lines to respect the original
                # input. Else, these unknown answers will be added as correct
                # answers.
                other_answers = [
                    answer
                    for answer in raw_question["answers"]
                    if answer.lower().strip() not in normalized_proposed_answers
                ]
                if has_open_choice:
                    open_choice = "\n".join(other_answers)
                else:
                    for answer in other_answers:
                        answers.append({"answer": answer.strip(), "is_correct": True})

            else:  # Linked
                answers = []
                error = False
                for answer in raw_question["proposed_answers"]:
                    if "-->" not in answer:
                        errored_questions.append((raw_question, "linked_answers"))
                        error = True
                        break
                    parts = answer.split("-->", 2)
                    if len(parts) != 2:
                        errored_questions.append((raw_question, "linked_answers"))
                        error = True
                        break
                    answers.append((parts[0].strip(), parts[1].strip()))
                if error:
                    continue

            # Tags
            # FIXME: multiple tags in different positions may have the same
            #  name. We should figure out if it's the case and warn the user.
            question_tags = []
            if raw_question["tags"]:
                for tag in raw_question["tags"]:
                    if tag.lower() in tags:
                        question_tags.append(tags[tag.lower()])
                    else:
                        tag_obj = Tag.objects.filter(name__iexact=tag).first()
                        if not tag_obj and form.cleaned_data["create_missing_tags"]:
                            # The tag is created as a root node. It can be moved afterward.
                            tag_obj = Tag(name=tag, parent=None)
                            tag_obj.save()

                        if tag_obj:
                            question_tags.append(tag_obj)
                            tags[tag.lower()] = tag_obj

            # Actual question creation
            if question_type == QUESTION_OPEN:
                question = Question.create_open(
                    question=raw_question["question"],
                    answer=answers,
                    locale=locale,
                    difficulty=difficulty,
                    illustration=None,
                    comment=raw_question["comment"],
                    tags=question_tags,
                    source=source,
                    user=self.request.user,
                )
            elif question_type == QUESTION_MCQ:
                question = Question.create_mcq(
                    question=raw_question["question"],
                    answers=answers,
                    has_open_answer=has_open_choice,
                    open_answer=open_choice,
                    locale=locale,
                    difficulty=difficulty,
                    illustration=None,
                    comment=raw_question["comment"],
                    tags=question_tags,
                    source=source,
                    user=self.request.user,
                )
            else:
                question = Question.create_linked(
                    question=raw_question["question"],
                    answers=answers,
                    locale=locale,
                    difficulty=difficulty,
                    illustration=None,
                    comment=raw_question["comment"],
                    tags=question_tags,
                    source=source,
                    user=self.request.user,
                )

            question.imported = True
            question.save()
            created_questions.append(question)

        return render(
            request=self.request,
            template_name="management/question-import-summary.html",
            context={
                "created_questions": created_questions,
                "created_questions_pks": ",".join(
                    [str(question.pk) for question in created_questions]
                ),
                "errored_questions": errored_questions,
                "locales": QuestionLocale.objects.all(),
            },
        )


class UndoImportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """
    This view deletes the given questions, but ensures that all questions are
    - created by the logged-in user;
    - created by an import.
    It then redirects to the import page.
    """

    permission_required = "quizz.add_question"

    @transaction.atomic
    def post(self, *args, **kwargs):
        pks = self.request.POST.get("pks", "").split(",")
        count = 0
        for pk in pks:
            try:
                question = Question.objects.filter(pk=int(pk)).first()
                if (
                    not question
                    or not question.imported
                    or question.creator != self.request.user
                ):
                    continue
                question.delete()
                count += 1
            except ValueError:
                pass

        messages.success(
            self.request,
            _(
                "The import was cancelled, removing %(count)d questions. You "
                "can retry by sending this form again with an updated "
                "spreadsheet."
            )
            % {"count": count},
        )

        return HttpResponseRedirect(reverse_lazy("quizz:management:questions-import"))
