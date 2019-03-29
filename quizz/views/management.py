from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.models import User
from django.db import transaction
from django.db.models import F, Count, Q, Sum, Max, Case, When, IntegerField, FloatField
from django.http import Http404, HttpResponseRedirect
from django.shortcuts import render, get_object_or_404
from django.urls import reverse_lazy, reverse
from django.utils.functional import cached_property
from django.utils.translation import gettext_lazy as _, ngettext_lazy
from django.views import View
from django.views.generic import ListView
from django.views.generic.edit import DeleteView, FormView
from openpyxl import load_workbook

from quizz.views.public import QuizzesListMixin
from ..forms.management import (
    ImportQuestionsForm,
    ManageQuizzLinkedAnswerFormSet,
    ManageQuizzQuestionForm,
    ManageQuizzSimpleAnswerFormSet,
    MigrateTagsForm,
)
from ..models.questions import Contest, Question, QuestionLocale, Tag
from quizz.models import QUESTION_OPEN, QUESTION_MCQ, QUESTION_LINKED, Quizz


class QuestionsListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    template_name = "management/question-list.html"
    context_object_name = "questions"

    permission_required = "quizz.view_question"

    model = Question
    paginate_by = 20

    @cached_property
    def sort_by(self):
        sort: str = self.request.GET.get("sort", "name")
        sort_reversed = sort.startswith("-")
        return sort.strip("-"), sort_reversed

    def get_context_data(self, **kwargs):
        context = super(QuestionsListView, self).get_context_data(**kwargs)

        context["sort_by"] = self.sort_by[0]
        context["sort_reversed"] = self.sort_by[1]

        return context

    def get_queryset(self):
        questions = (
            Question.objects.all()
            .annotate(
                user_answers_count=Count(
                    "user_answers",
                    filter=Q(user_answers__finished_at__isnull=False),
                    output_field=FloatField(),
                ),
                success_rate=Sum(
                    Case(
                        When(user_answers__success="PERFECT", then=1.0),
                        default=0.0,
                        output_field=FloatField(),
                    ),
                    output_field=FloatField(),
                )
                / F("user_answers_count")
                * 100,
            )
            .prefetch_related("tags")
            .select_related("locale", "source")
        )
        sort_by, sort_reversed = self.sort_by
        rev = "-" if sort_reversed else ""

        if sort_by == "difficulty":
            questions = questions.order_by(rev + "difficulty")
        elif sort_by == "source":
            questions = questions.order_by(
                F("source__name").asc(nulls_last=True, descending=sort_reversed)
            )
        elif sort_by == "illustration":
            # “Ordered” sorting for illustrations is with illustrations first,
            # so we need to have empty values last, corresponding to a reverse
            # sorting in SQL.
            questions = questions.order_by(
                ("" if sort_reversed else "-") + "illustration"
            )
        elif sort_by == "type":
            questions = questions.order_by(rev + "type")
        elif sort_by == "locale":
            questions = questions.order_by(rev + "locale__name")
        elif sort_by == "created":
            questions = questions.order_by(F("created_at").asc(nulls_last=True, descending=not sort_reversed))
        elif sort_by == "updated":
            questions = questions.order_by(F("updated_at").asc(nulls_last=True, descending=not sort_reversed))
        elif sort_by == "answered":
            questions = questions.order_by(
                ("" if sort_reversed else "-") + "user_answers_count"
            )
        elif sort_by == "success-rate":
            questions = questions.order_by(
                F("success_rate").asc(nulls_last=True, descending=not sort_reversed)
            )
        else:
            questions = questions.order_by(rev + "question")

        return questions


class QuestionsImportView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    template_name = "management/question-import.html"
    form_class = ImportQuestionsForm
    permission_required = "quizz.add_question"

    @transaction.atomic
    def form_valid(self, form):
        wb = load_workbook(self.request.FILES["spreadsheet"])
        ws = wb.active

        header_found = False
        values_started = False

        created_questions = []
        errored_questions = []

        locales = {}
        sources = {}
        tags = {}

        for row in ws.iter_rows(values_only=True):
            len_row = len(row)

            # We try to detect and skip an eventual header: the first row with
            # label-style values. If we miss it, it's not the end of the world
            # as such values will probably be invalid anyway (and users may still
            # delete the invalid question afterward).
            if not values_started and not header_found and row:
                normalized_row = [str(cell).lower() for cell in row]

                if (
                    (
                        len_row > 0
                        and (
                            "lang" in normalized_row[0]
                            or "language" in normalized_row[0]
                            or "locale" in normalized_row[0]
                        )
                    )
                    and (len_row > 1 and "question" in normalized_row[1])
                    and (
                        len_row > 2
                        and "proposition" in normalized_row[2]
                        or "proposed" in normalized_row[2]
                    )
                    and (
                        len_row > 3
                        and (
                            "answer" in normalized_row[3]
                            or "réponse" in normalized_row[3]
                            or "réponse" in normalized_row[3]
                        )
                    )
                    and (len_row > 4 and "comment" in normalized_row[4])
                    and (len_row > 5 and "tag" in normalized_row[5])
                    and (len_row > 6 and "origin" in normalized_row[6])
                    and (len_row > 7 and "difficult" in normalized_row[7])
                ):
                    header_found = True
                    values_started = True
                    continue

            if row:
                values_started = True

            # We store the raw content of the future question in a dict
            raw_question = {
                "locale": str(row[0]).strip() if len_row > 0 and row[0] else None,
                "question": str(row[1]).strip() if len_row > 1 and row[1] else None,
                "proposed_answers": str(row[2]).splitlines()
                if len_row > 2 and row[2]
                else None,
                "answers": str(row[3]).splitlines() if len_row > 3 and row[3] else None,
                "comment": str(row[4]).strip() if len_row > 4 and row[4] else None,
                "tags": str(row[5]).splitlines() if len_row > 5 and row[5] else None,
                "source": str(row[6]).strip() if len_row > 6 and row[6] else None,
                "difficulty": str(row[7]).strip() if len_row > 7 and row[7] else None,
            }

            # If the question is already obviously invalid (missing required
            # fields), we skip it.
            if not raw_question["question"] or (
                not raw_question["proposed_answers"] and not raw_question["answers"]
            ):
                errored_questions.append((raw_question, "missing_fields"))
                continue

            # Locale
            if raw_question["locale"].lower() in locales:
                locale = locales[raw_question["locale"].lower()]
            else:
                locale = QuestionLocale.objects.filter(
                    name__istartswith=raw_question["locale"]
                ).first()
                if not locale:
                    locale = QuestionLocale.objects.filter(
                        code__istartswith=raw_question["locale"]
                    ).first()
                if not locale:
                    errored_questions.append((raw_question, "locale"))
                    continue

                locales[raw_question["locale"].lower()] = locale

            # Difficulty
            if raw_question["difficulty"]:
                normalized_difficulty = raw_question["difficulty"].lower()
                if normalized_difficulty in ["1", "e", "easy"]:
                    difficulty = 1
                elif normalized_difficulty in ["2", "m", "medium"]:
                    difficulty = 2
                elif normalized_difficulty in ["3", "h", "hard"]:
                    difficulty = 3
                else:
                    difficulty = int(form.cleaned_data["default_difficulty"])
            else:
                difficulty = int(form.cleaned_data["default_difficulty"])

            # Source
            # We try to find a corresponding contest (skipping punctuation if needed);
            # if nothing can be found, we create a new one.
            source = None
            if raw_question["source"]:
                normalized_source = raw_question["source"].lower()
                if normalized_source in sources:
                    source = sources[normalized_source]
                else:
                    source = Contest.objects.filter(
                        name__iexact=normalized_source
                    ).first()
                    if not source:
                        no_pct_source = (
                            normalized_source.replace(",", "")
                            .replace(".", "")
                            .replace(";", "")
                            .replace(":", "")
                        )
                        source = Contest.objects.filter(
                            name__iexact=no_pct_source
                        ).first()
                        if not source:
                            source = Contest(name=raw_question["source"])
                            source.save()
                    sources[normalized_source] = source

            # Question type
            if not raw_question["proposed_answers"]:
                question_type = QUESTION_OPEN
            elif not raw_question["answers"]:
                question_type = (
                    QUESTION_LINKED
                    if all(
                        ["-->" in answer for answer in raw_question["proposed_answers"]]
                    )
                    else QUESTION_MCQ
                )
            else:
                question_type = QUESTION_MCQ

            # Answers
            has_open_choice = False
            open_choice = None

            if question_type == QUESTION_OPEN:
                answers = "\n".join(
                    [answer.strip() for answer in raw_question["answers"]]
                )

            elif question_type == QUESTION_MCQ:
                answers = []

                normalized_proposed_answers = [
                    answer.lower().strip()
                    for answer in raw_question["proposed_answers"]
                ]
                normalized_answers = [
                    answer.lower().strip() for answer in raw_question["answers"]
                ]

                for answer in raw_question["proposed_answers"]:
                    if answer.lower() in [
                        "other:",
                        "other :",
                        "autre:",
                        "autre :",  # space [SP]
                        "autre :",  # no-break space [NBSP]
                        "autre :",  # narrow no-break space [NNBSP]
                        "autres:",
                        "autres :",  # space [SP]
                        "autres :",  # no-break space [NBSP]
                        "autres :",  # narrow no-break space [NNBSP]
                    ]:
                        has_open_choice = True
                    else:
                        answers.append(
                            {
                                "answer": answer.strip(),
                                "is_correct": answer.lower().strip() in normalized_answers,
                            }
                        )

                # We process unknown answers here (i.e. answers in the “Answers”
                # list but not in the “Proposed answers” one).
                # If we have an open answer, these will be the answer, and they
                # will be concatenated with new lines to respect the original
                # input. Else, these unknown answers will be added as correct
                # answers.
                other_answers = [
                    answer
                    for answer in raw_question["answers"]
                    if answer.lower().strip() not in normalized_proposed_answers
                ]
                if has_open_choice:
                    open_choice = "\n".join(other_answers)
                else:
                    for answer in other_answers:
                        answers.append({"answer": answer.strip(), "is_correct": True})

            else:  # Linked
                answers = []
                error = False
                for answer in raw_question["proposed_answers"]:
                    if "-->" not in answer:
                        errored_questions.append((raw_question, "linked_answers"))
                        error = True
                        break
                    parts = answer.split("-->", 2)
                    if len(parts) != 2:
                        errored_questions.append((raw_question, "linked_answers"))
                        error = True
                        break
                    answers.append((parts[0].strip(), parts[1].strip()))
                if error:
                    continue

            # Tags
            # FIXME: multiple tags in different positions may have the same
            #  name. We should figure out if it's the case and warn the user.
            question_tags = []
            if raw_question["tags"]:
                for tag in raw_question["tags"]:
                    if tag.lower() in tags:
                        question_tags.append(tags[tag.lower()])
                    else:
                        tag_obj = Tag.objects.filter(name__iexact=tag).first()
                        if not tag_obj and form.cleaned_data["create_missing_tags"]:
                            # The tag is created as a root node. It can be moved afterward.
                            tag_obj = Tag(name=tag, parent=None)
                            tag_obj.save()

                        if tag_obj:
                            question_tags.append(tag_obj)
                            tags[tag.lower()] = tag_obj

            # Actual question creation
            if question_type == QUESTION_OPEN:
                question = Question.create_open(
                    question=raw_question["question"],
                    answer=answers,
                    locale=locale,
                    difficulty=difficulty,
                    illustration=None,
                    comment=raw_question["comment"],
                    tags=question_tags,
                    source=source,
                    user=self.request.user,
                )
            elif question_type == QUESTION_MCQ:
                question = Question.create_mcq(
                    question=raw_question["question"],
                    answers=answers,
                    has_open_answer=has_open_choice,
                    open_answer=open_choice,
                    locale=locale,
                    difficulty=difficulty,
                    illustration=None,
                    comment=raw_question["comment"],
                    tags=question_tags,
                    source=source,
                    user=self.request.user,
                )
            else:
                question = Question.create_linked(
                    question=raw_question["question"],
                    answers=answers,
                    locale=locale,
                    difficulty=difficulty,
                    illustration=None,
                    comment=raw_question["comment"],
                    tags=question_tags,
                    source=source,
                    user=self.request.user,
                )

            question.imported = True
            question.save()
            created_questions.append(question)

        return render(
            request=self.request,
            template_name="management/question-import-summary.html",
            context={
                "created_questions": created_questions,
                "created_questions_pks": ",".join(
                    [str(question.pk) for question in created_questions]
                ),
                "errored_questions": errored_questions,
                "locales": QuestionLocale.objects.all(),
            },
        )


class UndoImportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """
    This view deletes the given questions, but ensures that all questions are
    - created by the logged-in user;
    - created by an import.
    It then redirects to the import page.
    """

    permission_required = "quizz.add_question"

    @transaction.atomic
    def post(self, *args, **kwargs):
        pks = self.request.POST.get("pks", "").split(",")
        count = 0
        for pk in pks:
            try:
                question = Question.objects.filter(pk=int(pk)).first()
                if (
                    not question
                    or not question.imported
                    or question.creator != self.request.user
                ):
                    continue
                question.delete()
                count += 1
            except ValueError:
                pass

        messages.success(
            self.request,
            _(
                "The import was cancelled, removing %(count)d questions. You "
                "can retry by sending this form again with an updated "
                "spreadsheet."
            )
            % {"count": count},
        )

        return HttpResponseRedirect(reverse_lazy("quizz:management:questions-import"))


class EditQuestionView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    template_name = "management/question-edit.html"
    form_class = ManageQuizzQuestionForm

    def get_permission_required(self):
        if self.edited_question:
            return ("quizz.change_question",)
        else:
            return ("quizz.add_question",)

    @cached_property
    def edited_question(self):
        """
        Returns the currently edited question, or None if this form is used for creation.
        :rtype: Question
        """
        if "pk" in self.kwargs:
            question = Question.objects.filter(pk=self.kwargs["pk"]).first()
            if not question:
                raise Http404()
            return question
        else:
            return None

    def get_form(self, form_class=None):
        """
        If we're editing a question, disables the question type field,
        as types cannot be changed.
        Even if the user try to temper with that, any new value will be ignored.
        """
        form = super(EditQuestionView, self).get_form(form_class=form_class)

        if self.edited_question:
            form.fields["type"].disabled = True

        return form

    def get_initial(self):
        if not self.edited_question:
            return {}

        return {
            "type": self.edited_question.type,
            "locale": self.edited_question.locale,
            "difficulty": self.edited_question.difficulty,
            "illustration": self.edited_question.illustration,
            "question": self.edited_question.question,
            "answer_comment": self.edited_question.answer_comment,
            "open_answer": self.edited_question.open_valid_answer,
            "has_open_choice": self.edited_question.has_open_choice,
            "source": self.edited_question.source,
            "tags": self.edited_question.tags.all(),
        }

    def get_initial_answers_formset(self):
        if not self.edited_question or self.edited_question.type != QUESTION_MCQ:
            return []

        return [
            {"answer": answer.answer, "is_correct": answer.is_correct}
            for answer in self.edited_question.answers.filter(is_deleted=False)
        ]

    def get_initial_linked_answers_formset(self):
        if not self.edited_question or self.edited_question.type != QUESTION_LINKED:
            return []

        return [
            {"answer": answer.answer, "linked_answer": answer.linked_answer.answer}
            for answer in self.edited_question.answers.filter(
                is_deleted=False
            ).prefetch_related("linked_answer")
        ]

    def get_context_data(self, form_answers=None, form_linked_answers=None, **kwargs):
        context = super(EditQuestionView, self).get_context_data(**kwargs)

        context["question"] = self.edited_question
        context["form_answers"] = form_answers or ManageQuizzSimpleAnswerFormSet(
            prefix="answers", initial=self.get_initial_answers_formset()
        )
        context["form_linked_answers"] = (
            form_linked_answers
            or ManageQuizzLinkedAnswerFormSet(
                prefix="linked-answers",
                initial=self.get_initial_linked_answers_formset(),
            )
        )

        return context

    def form_invalid(self, form, form_answers=None, form_linked_answers=None, **kwargs):
        return self.render_to_response(
            self.get_context_data(
                form=form,
                form_answers=form_answers,
                form_linked_answers=form_linked_answers,
            )
        )

    def form_valid(self, form):
        question = self.edited_question
        is_update = question is not None

        # The question type can't be changed. It would be too difficult/touchy
        # to keep history clean for past quizz.
        question_type = question.type if question else form.cleaned_data["type"]
        form_answers = None
        form_linked_answers = None

        if question_type == QUESTION_MCQ:
            form_answers = ManageQuizzSimpleAnswerFormSet(
                self.request.POST, self.request.FILES, prefix="answers"
            )
        elif question_type == QUESTION_LINKED:
            form_linked_answers = ManageQuizzLinkedAnswerFormSet(
                self.request.POST, self.request.FILES, prefix="linked-answers"
            )

        if (
            form.has_changed()
            or (not form_answers or form_answers.has_changed())
            or (not form_linked_answers or form_linked_answers.has_changed())
        ) and self.request.user.is_authenticated:
            user = self.request.user
        else:
            user = None

        if (not form_answers or form_answers.is_valid()) and (
            not form_linked_answers or form_linked_answers.is_valid()
        ):
            if question_type == QUESTION_OPEN:
                question = Question.create_or_update_open(
                    instance=question,
                    question=form.cleaned_data["question"],
                    answer=form.cleaned_data["open_answer"],
                    locale=form.cleaned_data["locale"],
                    difficulty=form.cleaned_data["difficulty"],
                    illustration=self.request.FILES.get("illustration", None),
                    delete_illustration=form.cleaned_data["delete_illustration"]
                    if "delete_illustration" in form.cleaned_data
                    else False,
                    comment=form.cleaned_data["answer_comment"],
                    tags=form.cleaned_data["tags"],
                    source=form.cleaned_data["source"],
                    user=user,
                )

            elif question_type == QUESTION_MCQ:
                answers = []
                if form_answers.has_changed():
                    for form_answer in form_answers.forms:
                        if (
                            "answer" in form_answer.cleaned_data
                            and "is_correct" in form_answer.cleaned_data
                            and form_answer.cleaned_data["answer"].strip()
                        ):
                            answers.append(
                                {
                                    "answer": form_answer.cleaned_data[
                                        "answer"
                                    ].strip(),
                                    "is_correct": form_answer.cleaned_data[
                                        "is_correct"
                                    ],
                                }
                            )

                question = Question.create_or_update_mcq(
                    instance=question,
                    question=form.cleaned_data["question"],
                    answers=answers,
                    has_open_answer=form.cleaned_data["has_open_choice"],
                    open_answer=form.cleaned_data["open_answer"],
                    locale=form.cleaned_data["locale"],
                    difficulty=form.cleaned_data["difficulty"],
                    illustration=self.request.FILES.get("illustration", None),
                    delete_illustration=form.cleaned_data["delete_illustration"]
                    if "delete_illustration" in form.cleaned_data
                    else False,
                    comment=form.cleaned_data["answer_comment"],
                    tags=form.cleaned_data["tags"],
                    source=form.cleaned_data["source"],
                    user=user,
                )

                if not any([answer["is_correct"] for answer in answers]) and (
                    not question.has_open_choice or not question.open_valid_answer
                ):
                    messages.warning(
                        self.request,
                        _(
                            "All answers are marked “incorrect” in the question “%(question)s”, and there is no "
                            "valid open choice. The question will only be considered valid if submitted blank. "
                            "Make sure that's indeed what you want."
                        )
                        % {"question": form.cleaned_data["question"]},
                    )

            elif question_type == QUESTION_LINKED:
                answers = []
                if form_linked_answers.has_changed():
                    for form_linked_answer in form_linked_answers.forms:
                        if (
                            "answer" in form_linked_answer.cleaned_data
                            and form_linked_answer.cleaned_data["answer"].strip()
                            and "linked_answer" in form_linked_answer.cleaned_data
                            and form_linked_answer.cleaned_data["linked_answer"].strip()
                        ):
                            answers.append(
                                (
                                    form_linked_answer.cleaned_data["answer"].strip(),
                                    form_linked_answer.cleaned_data[
                                        "linked_answer"
                                    ].strip(),
                                )
                            )

                question = Question.create_or_update_linked(
                    instance=question,
                    answers=answers,
                    question=form.cleaned_data["question"],
                    locale=form.cleaned_data["locale"],
                    difficulty=form.cleaned_data["difficulty"],
                    illustration=self.request.FILES.get("illustration", None),
                    delete_illustration=form.cleaned_data["delete_illustration"]
                    if "delete_illustration" in form.cleaned_data
                    else False,
                    comment=form.cleaned_data["answer_comment"],
                    tags=form.cleaned_data["tags"],
                    source=form.cleaned_data["source"],
                    user=user,
                )

            if is_update:
                messages.success(
                    self.request,
                    _(
                        "Question “%(question)s” (%(question_type)s) updated successfully."
                    )
                    % {
                        "question": question.question,
                        "question_type": question.verbose_type,
                    },
                )
            else:
                messages.success(
                    self.request,
                    _(
                        "Question “%(question)s” (%(question_type)s) created successfully."
                    )
                    % {
                        "question": question.question,
                        "question_type": question.verbose_type,
                    },
                )

            if (
                "and_after" in self.request.POST
                and self.request.POST["and_after"] == "ANOTHER"
            ):
                return HttpResponseRedirect(
                    reverse_lazy("quizz:management:questions-create")
                )
            else:
                return HttpResponseRedirect(reverse_lazy("quizz:management:questions"))

        else:
            return self.form_invalid(
                form=form,
                form_answers=form_answers,
                form_linked_answers=form_linked_answers,
            )


class QuestionDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    template_name = "management/question-delete.html"
    context_object_name = "question"

    permission_required = "quizz.delete_question"

    model = Question

    def delete(self, *args, **kwargs):
        question: Question = self.get_object()

        question.delete()

        messages.success(
            self.request,
            _("Question “%(question)s” deleted successfully.")
            % {"question": question.question},
        )

        return HttpResponseRedirect(reverse_lazy("quizz:management:questions"))


class MigrateTagsView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    template_name = "management/tags-migrate.html"
    permission_required = ("quizz.delete_tag", "quizz.change_question")

    form_class = MigrateTagsForm
    success_url = reverse_lazy("quizz:management:tags-migration")

    def form_valid(self, form):
        old_tag: Tag = form.cleaned_data["old_tag"]
        target_tag: Tag = form.cleaned_data["new_tag"]
        delete_old = form.cleaned_data["delete_old_tag"]

        if old_tag == target_tag:
            messages.error(self.request, _("You cannot migrate a tag to itself"))
            return HttpResponseRedirect(self.get_success_url())

        count = 0
        for question in Question.objects.filter(tags=old_tag):
            question.tags.add(target_tag)
            if not delete_old:
                question.tags.remove(old_tag)
            count += 1

        if delete_old:
            # If the tag have children, we want to move them to its parent
            # because deleting a tag deletes all its children
            has_children = not old_tag.is_leaf_node()
            if has_children:
                parent = old_tag.parent
                for child in old_tag.get_children():
                    child.parent = parent
                    child.save()

            old_tag.delete()

            if has_children:
                Tag.objects.rebuild()

            messages.success(
                self.request,
                ngettext_lazy(
                    "The tag %(old_tag)s was deleted, and %(questions_count)d "
                    "question was moved to the tag %(new_tag)s.",
                    "The tag %(old_tag)s was deleted, and %(questions_count)d "
                    "questions were moved to the tag %(new_tag)s.",
                    count,
                )
                % {
                    "old_tag": old_tag.name,
                    "new_tag": target_tag.name,
                    "questions_count": count,
                },
            )
        else:
            messages.success(
                self.request,
                ngettext_lazy(
                    "%(questions_count)d question was moved to the tag %(new_tag)s.",
                    "%(questions_count)d questions were moved to the tag %(new_tag)s.",
                    count,
                )
                % {"new_tag": target_tag.name, "questions_count": count},
            )

        return HttpResponseRedirect(self.get_success_url())


class QuizzesListView(LoginRequiredMixin, PermissionRequiredMixin, QuizzesListMixin):
    permission_required = "quizz.view_quizz"

    def get_base_queryset(self):
        return Quizz.objects.all()

    def get_context_data(self, *args, **kwargs):
        context = super(QuizzesListView, self).get_context_data(*args, **kwargs)

        context["management"] = True
        context["quizzes_user"] = None

        return context


class UserProfileView(LoginRequiredMixin, PermissionRequiredMixin, QuizzesListMixin):
    permission_required = ("quizz.view_quizz", "auth.view_user")

    @cached_property
    def user(self):
        return get_object_or_404(User, username=self.kwargs["username"])

    def get_base_queryset(self):
        return Quizz.objects.filter(user=self.user)

    def get_context_data(self, *args, **kwargs):
        context = super(UserProfileView, self).get_context_data(*args, **kwargs)

        context["management"] = True
        context["quizzes_user"] = self.user

        return context


class UsersListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    template_name = "management/users-list.html"
    context_object_name = "users"

    permission_required = "auth.view_user"

    model = User
    paginate_by = 20

    def get_queryset(self):
        return (
            super(UsersListView, self)
            .get_queryset()
            .annotate(
                quizzes_count=Count("quizzes", distinct=True),
                questions_count=Count("quizzes__questions", distinct=True),
                groups_count=Max(
                    Case(
                        When(groups__isnull=False, then=1),
                        default=0,
                        output_field=IntegerField(),
                    )
                ),
            )
            .order_by(
                "-is_superuser",
                "-is_staff",
                "-groups_count",
                "last_name",
                "first_name",
                "username",
            )
            .select_related("profile")
            .prefetch_related("groups")
        )
